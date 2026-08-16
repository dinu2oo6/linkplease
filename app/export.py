"""Export the run as a spreadsheet.

A real .xlsx with four sheets when openpyxl is available, CSV otherwise. Both
open in Excel; the workbook just keeps the checkpoints, the DM log, the blocked
duplicates and the summary on separate tabs instead of concatenating them.

Everything here is read straight from the ledger at request time, so the export
and the graded numbers cannot disagree.
"""
import csv
import io
import json
import time

from . import checkpoints, db, stats


def _summary_rows() -> list[list]:
    core = stats.core_stats()
    detail = stats.verbose_stats()["detail"]
    return [
        ["Metric", "Value", "Meaning"],
        ["Sent", core["sent"], "DMs the API confirmed as delivered"],
        ["Queued", core["queued"], "Still owed, incl. awaiting confirmation"],
        ["Failed", core["failed"], "Gave up after retries"],
        ["Duplicates blocked", core["duplicates_blocked"], "DMs correctly not sent"],
        [],
        ["Distinct events received", detail["events_distinct"], ""],
        ["Total deliveries", detail["event_deliveries"], "Includes redeliveries"],
        ["Redeliveries", detail["events_redelivered"], "Same event sent more than once"],
        ["Cancelled by delete", detail["cancelled_by_delete"], ""],
        ["Suppressed (delete first)", detail["suppressed_deleted"], ""],
        ["Total send requests", detail["sends_total"], "Includes retries and resends"],
        ["Resends after failed delivery", detail["resends_issued"], ""],
        ["Rules configured", detail["rules"], ""],
        [],
        ["Exported at", time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()), ""],
    ]


def _checkpoint_rows() -> list[list]:
    data = checkpoints.all_checkpoints()
    rows = [["Checkpoint", "Requirement", "Result", "Detail", "Evidence"]]
    for c in data["checkpoints"]:
        rows.append([c["title"], c["requirement"], c["state"].upper(), c["detail"],
                     json.dumps(c["evidence"])])
    s = data["summary"]
    rows += [[], ["TOTAL", f"{s['total']} checkpoints",
                  f"{s['pass']} pass / {s['fail']} fail / {s['pending']} pending", "", ""]]
    return rows


def _dm_rows() -> list[list]:
    rows = [["Username", "User ID", "Comment", "Keyword", "DM message", "State",
             "Attempts", "Resends", "DM ID", "Seconds to resolve", "Error"]]
    for r in db.query(
        """SELECT COALESCE(t.username, a.username) AS username, t.user_id,
                  c.text AS comment, ru.keyword, t.message, t.state, t.attempts,
                  t.resend_count, t.dm_id, t.created_at, t.updated_at, t.last_error
           FROM dm_tasks t
           LEFT JOIN demo_accounts a ON a.user_id = t.user_id
           LEFT JOIN comments c      ON c.comment_id = t.comment_id
           LEFT JOIN rules ru        ON ru.rule_id = t.rule_id
           ORDER BY t.created_at"""
    ):
        rows.append([
            r["username"] or r["user_id"], r["user_id"], r["comment"], r["keyword"],
            r["message"], r["state"], r["attempts"], r["resend_count"], r["dm_id"],
            round((r["updated_at"] or 0) - (r["created_at"] or 0), 1), r["last_error"],
        ])
    return rows


def _blocked_rows() -> list[list]:
    rows = [["User ID", "Decision", "Rule", "When", "Why"]]
    for r in db.query(
        """SELECT m.user_id, m.decision, ru.keyword, m.created_at
           FROM match_decisions m LEFT JOIN rules ru ON ru.rule_id = m.rule_id
           WHERE m.decision <> ? ORDER BY m.created_at""", (db.D_CREATED,)
    ):
        why = ("already owed this DM for this rule" if r["decision"] == db.D_DUPLICATE
               else "comment was already deleted")
        rows.append([r["user_id"], r["decision"], r["keyword"],
                     time.strftime("%H:%M:%S", time.gmtime(r["created_at"])), why])
    return rows


SHEETS = [
    ("Summary", _summary_rows),
    ("Checkpoints", _checkpoint_rows),
    ("DM log", _dm_rows),
    ("Blocked duplicates", _blocked_rows),
]


def to_csv() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for name, builder in SHEETS:
        writer.writerow([f"=== {name} ==="])
        writer.writerows(builder())
        writer.writerow([])
    return buf.getvalue()


def to_xlsx() -> bytes | None:
    """Real workbook if openpyxl is installed, else None so the caller falls back."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    colours = {"PASS": "C6EFCE", "FAIL": "FFC7CE", "PENDING": "FFEB9C"}

    for name, builder in SHEETS:
        ws = wb.create_sheet(name[:31])
        for row in builder():
            ws.append(row if row else [""])
        for cell in ws[1]:
            cell.fill, cell.font = head_fill, head_font
        if name == "Checkpoints":
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if cell.value in colours:
                        cell.fill = PatternFill("solid", fgColor=colours[cell.value])
                        cell.font = Font(bold=True)
        widths = {"A": 26, "B": 40, "C": 14, "D": 60, "E": 44}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
